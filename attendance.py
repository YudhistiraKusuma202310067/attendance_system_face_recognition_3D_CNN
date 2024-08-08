import cv2
from tkinter import *
from tkinter import ttk
import tkinter.messagebox as messagebox
from PIL import Image, ImageTk
from mtcnn import MTCNN
from keras.models import load_model
import numpy as np
import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
import matplotlib.pyplot as plt
from Silent_Face_Anti_Spoofing_master.test import test

# Load trained model
model = load_model('model/test dataset8.h5')
dataset_dir = 'be_augmented_dataset_5'
labels = [folder.replace('_', ' ') for folder in os.listdir(dataset_dir) if os.path.isdir(os.path.join(dataset_dir, folder))]

# Load MTCNN for face detection
detector = MTCNN()

# Function to end attendance and close the application
def end_attendance():
    cap.release()
    cv2.destroyAllWindows()
    root.destroy()

# Buat fungsi handler event terpisah
def handle_status_click(event):
    item_id = student_table.identify_row(event.y)  # Dapatkan id baris yang diklik
    if item_id:  # Pastikan id baris ada
        current_item = student_table.item(item_id)
        if current_item:
            number = current_item['values'][0]
            current_status = current_item['values'][3]  # Ambil status saat ini dari baris yang dipilih
            open_status_window(number, current_status)  # Panggil fungsi untuk membuka jendela status

# Fungsi untuk membuka jendela pop-up ketika status pada tabel ditekan
def open_status_window(number, current_status):
    # Fungsi untuk menyimpan perubahan status
    def save_status():
        # Perbarui status pada tabel
        student_table.item(student_table.selection(), values=(number, student_table.item(student_table.selection(), 'values')[1],
                                                            student_table.item(student_table.selection(), 'values')[2],
                                                            new_status_var.get(),
                                                            student_table.item(student_table.selection(), 'values')[4]))
        # Tutup jendela pop-up
        status_window.destroy()

    # Buat jendela pop-up baru
    status_window = Toplevel(root)
    status_window.title("Status")
    
    # Atur lebar jendela
    status_window_width = 260
    status_window_height = 140
    screen_width = status_window.winfo_screenwidth()
    screen_height = status_window.winfo_screenheight()
    x_coordinate = (screen_width / 2) - (status_window_width / 2)
    y_coordinate = (screen_height / 2) - (status_window_height / 2)
    status_window.geometry("%dx%d+%d+%d" % (status_window_width, status_window_height, x_coordinate, y_coordinate))

    # Buat StringVar untuk menyimpan status baru
    new_status_var = StringVar()
    new_status_var.set(current_status)  # Set nilai default dengan status saat ini

    # Buat radio button untuk setiap opsi status
    radio_button_hadir = ttk.Radiobutton(status_window, text="Hadir", variable=new_status_var, value="Hadir", command=save_status)
    radio_button_izin = ttk.Radiobutton(status_window, text="Izin", variable=new_status_var, value="Izin", command=save_status)
    radio_button_sakit = ttk.Radiobutton(status_window, text="Sakit", variable=new_status_var, value="Sakit", command=save_status)
    radio_button_alpha = ttk.Radiobutton(status_window, text="Alpha", variable=new_status_var, value="Alpha", command=save_status)
    radio_button_dispen = ttk.Radiobutton(status_window, text="Dispensasi", variable=new_status_var, value="Dispensasi", command=save_status)

    # Susun radio button dan tombol "Save" dalam jendela pop-up
    radio_button_hadir.grid(row=0, column=0, padx=10, sticky="w")
    radio_button_izin.grid(row=1, column=0, padx=10, sticky="w")
    radio_button_sakit.grid(row=2, column=0, padx=10, sticky="w")
    radio_button_alpha.grid(row=3, column=0, padx=10, sticky="w")
    radio_button_dispen.grid(row=4, column=0, padx=10, sticky="w")

# Update fungsi update_students untuk menambahkan event handler ke status pada tabel
def update_students(event=None):
    selected_class = class_var.get()
    students = class_data[selected_class]["student"]
    clear_table()
    
    if students:  # Jika ada mahasiswa dalam kelas
        for i, student in enumerate(students):
            student_npm = student["npm"]
            student_name = student["name"]
            student_status = "Alpha"
            
            # Insert student data into the table, including a separate column for status
            student_data = (i + 1, student_npm, student_name, student_status, "-")
            student_table.insert("", "end", values=student_data)

        # Bind event handler untuk status pada tabel
        student_table.bind('<Button-1>', handle_status_click)

    # Panggil load_and_display_data() di sini
    load_and_display_data()

# Function to clear student table
def clear_table():
    student_table.delete(*student_table.get_children())

# Function to update status and timestamp based on face recognition
def update_status_and_timestamp(student_name, timestamp_start):
    for row_id in student_table.get_children():
        student = student_table.item(row_id)["values"]
        if student[2] == student_name:
            if student[4] == "-":  # Check if timestamp is empty
                student_table.item(row_id, values=(student[0], student[1], student[2], "Hadir", datetime.now().strftime("%d %b %Y %H:%M:%S")))
                timestamp_end =  datetime.now()
                detected_face = (timestamp_end - timestamp_start).total_seconds() * 1000
                print("Waktu Deteksi Wajah: ", detected_face)
            break

# Function to load data from JSON and display it
def load_and_display_data():
    global alasan_entry, evaluasi_entry
    selected_class = class_var.get()
    
    # Load data from JSON
    class_info = class_data[selected_class]

    # Clear information frame before displaying new data
    clear_info_frame()

    # Display class information
    kode_mk_label = Label(info_frame, text="Kode MK")
    kode_mk_label.grid(row=0, column=0, sticky="w")
    kode_mk_value = Label(info_frame, width=50, anchor="w", text=": " + class_info["Kode_MK"])
    kode_mk_value.grid(row=0, column=1, sticky="w")

    nama_mk_label = Label(info_frame, text="Nama MK")
    nama_mk_label.grid(row=0, column=2, sticky="w")
    nama_mk_value = Label(info_frame, width=50, anchor="w", text=": " + selected_class)
    nama_mk_value.grid(row=0, column=3, sticky="w")

    kelas_label = Label(info_frame, text="Kelas")
    kelas_label.grid(row=1, column=0, sticky="w")
    kelas_value = Label(info_frame, text=": " + class_info["Kelas"])
    kelas_value.grid(row=1, column=1, sticky="w")

    nama_kelas_label = Label(info_frame, text="Nama Kelas")
    nama_kelas_label.grid(row=1, column=2, sticky="w")
    nama_kelas_label = Label(info_frame, text=": " + class_info["Kelas"])
    nama_kelas_label.grid(row=1, column=3, sticky="w")

    kode_label = Label(info_frame, text="Kode")
    kode_label.grid(row=2, column=0, sticky="w")
    kode_value = Label(info_frame, text=": " + class_info["Kode"])
    kode_value.grid(row=2, column=1, sticky="w")

    nama_dosen_label = Label(info_frame, text="Nama Dosen")
    nama_dosen_label.grid(row=2, column=2, sticky="w")
    nama_dosen_value = Label(info_frame, text=": " + class_info["Nama_Dosen"])
    nama_dosen_value.grid(row=2, column=3, sticky="w")

    no_urut = Label(info_frame, text="No. Urut")
    no_urut.grid(row=3, column=0, sticky="w")
    no_urut = Label(info_frame, text=": " + class_info["No_Urut"])
    no_urut.grid(row=3, column=1, sticky="w")

    pertemuan_ke_label = Label(info_frame, text="Pertemuan Ke")
    pertemuan_ke_label.grid(row=3, column=2, sticky="w")
    pertemuan_ke_value = Label(info_frame, text=": " + class_info["Pertemuan_Ke"])
    pertemuan_ke_value.grid(row=3, column=3, sticky="w")

    tanggal_label = Label(info_frame, text="Tanggal")
    tanggal_label.grid(row=4, column=0, sticky="w")
    tanggal_value = Label(info_frame, text=": " + datetime.now().strftime("%d-%m-%Y"))
    tanggal_value.grid(row=4, column=1, sticky="w")

    hari_label = Label(info_frame, text="Hari")
    hari_label.grid(row=4, column=2, sticky="w")
    hari_value = Label(info_frame, text=": " + class_info["Hari"])
    hari_value.grid(row=4, column=3, sticky="w")

    waktu_label = Label(info_frame, text="Waktu")
    waktu_label.grid(row=5, column=0, sticky="w")
    waktu_value = Label(info_frame, text=": " + class_info["Waktu"])
    waktu_value.grid(row=5, column=1, sticky="w")

    ruang_label = Label(info_frame, text="Ruang")
    ruang_label.grid(row=5, column=2, sticky="w")
    ruang_value = Label(info_frame, text=": " + class_info["Ruang"])
    ruang_value.grid(row=5, column=3, sticky="w")

    dosen_pengajar_label = Label(info_frame, text="Dosen Pengajar")
    dosen_pengajar_label.grid(row=6, column=0, sticky="w")
    dosen_pengajar_value = Label(info_frame, text=": " + class_info["Kode"])
    dosen_pengajar_value.grid(row=6, column=1, sticky="w")

    nama_label = Label(info_frame, text="Nama")
    nama_label.grid(row=6, column=2, sticky="w")
    nama_value = Label(info_frame, text=": " + class_info["Nama_Dosen"])
    nama_value.grid(row=6, column=3, sticky="w")

    rps_label = Label(info_frame, text="RPS")
    rps_label.grid(row=7, column=0, sticky="w")
    rps_value = Label(info_frame, text=": " + class_info["RPS"])
    rps_value.grid(row=7, column=1, sticky="w")

    # Create entry for "Alasan" and "Evaluasi Pembelajaran"
    alasan_label = Label(info_frame2, text="Berikan alasan (Jika tidak sesuai) ")
    alasan_label.grid(row=8, column=0, sticky="w")
    alasan_entry = Entry(info_frame2, width=95)
    alasan_entry.grid(row=8, column=1, sticky="w")

    evaluasi_label = Label(info_frame2, text="Evaluasi Pembelajaran (quiz, tugas, ujian) ")
    evaluasi_label.grid(row=9, column=0, sticky="w")
    evaluasi_entry = Entry(info_frame2, width=95)
    evaluasi_entry.grid(row=9, column=1, sticky="w")

# Function to clear information frame
def clear_info_frame():
    for widget in info_frame.winfo_children():
        widget.destroy()

# Create Tkinter window
root = Tk()
root.title("Attendance")

# Access webcam
cap = cv2.VideoCapture(0)

# Load class data from JSON
with open("student_class.json", "r") as f:
    class_data = json.load(f)

# Create a frame to hold the webcam feed
frame_label = Label(root)
frame_label.pack(side="right", padx=5)  # Place frame on the right side

# Create dropdown for class selection
class_var = StringVar()
class_var.set("Select Class")  # Set default value to placeholder
class_dropdown = ttk.Combobox(root, textvariable=class_var, state="readonly")
# class_dropdown["values"] = ["Select Class"] + list(class_data.keys())
class_dropdown["values"] = list(class_data.keys())
class_dropdown.pack(side="top", anchor="e", padx=10, pady=10)  # Place dropdown on top
class_dropdown.bind("<<ComboboxSelected>>", update_students)  # Bind function to event

# Create frame for additional information
info_frame = Frame(root)
info_frame.pack(side="top", padx=10, pady=10, anchor="w")

# Create frame for additional information
info_frame2 = Frame(root)
info_frame2.pack(side="top", padx=10, pady=10, anchor="w")

# Call the function to load and display data
load_and_display_data()

# Create student table
student_table = ttk.Treeview(root, columns=("no", "npm", "name", "status", "timestamp"), show="headings")
student_table.heading("no", text="No.")
student_table.column("no", width=50, anchor="center")
student_table.heading("npm", text="NPM")
student_table.heading("name", text="Name")
student_table.heading("status", text="Status")
student_table.heading("timestamp", text="Timestamp")
student_table.pack(side="top", padx=10)  # Place table below dropdown

# Set initial value for attendance_running
attendance_running = False

# Deklarasi variabel global
no_faces_count = 0
end_attendance_status = False

# Function to update the webcam feed
def update_frame():
    ret, frame = cap.read()
    if ret:
        global update_id, no_faces_count, end_attendance_status

        # Detect faces using MTCNN
        faces = detector.detect_faces(frame)

        if faces:
            no_faces_count = 0
            for face in faces:
                #Anti Spoof
                x, y, w, h = face['box']
                top_limit = max(0, y - 75)
                bottom_limit = min(frame.shape[0], y + h + 75)
                left_limit = max(0, x - 75)
                right_limit = min(frame.shape[1], x + w + 75)
                    
                # Crop area dari rambut hingga leher dengan area yang diperluas
                face_img_spoof = frame[top_limit:bottom_limit, left_limit:right_limit]

                # plt.imshow(face_img_spoof)
                # plt.axis('off')  # Turn off axis
                # plt.show()

                label = test(image=face_img_spoof,
                     model_dir="Silent_Face_Anti_Spoofing_master/resources/anti_spoof_models",
                     device_id=0)

                #Predict
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

                x, y, w, h = face['box']
                face_img = gray[y:y+h, x:x+w]
                face_img = cv2.resize(face_img, (50, 50))
                face_img = face_img.reshape(1, 50, 50, 1)  # Reshape sesuai dengan dimensi model

                # Recognize face
                if label == 1:
                    timestamp_start = datetime.now()
                    result = model.predict(face_img)
                    idx = result.argmax(axis=1)
                    confidence = result.max(axis=1) * 100
                    predicted_as = "N/A"
                    for i in idx:
                        if confidence > 80:
                            predicted_as = "%s (%.2f %%)" % (labels[i], confidence)
                            update_status_and_timestamp(labels[i], timestamp_start)
                        else:
                            predicted_as = "N/A"

                        print("Predicted as :", labels[i], "Persentase:", confidence)
                else:
                    predicted_as = "Fake Face Detected!!!"

                # Draw rectangle around face
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 0, 255), 2)

                # Display label and confidence
                cv2.putText(frame, predicted_as, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
        else:
            # Faces not detected, increment no_faces_count
            no_faces_count += 1
            if no_faces_count >= 10:
                end_attendance_status = True

        # Convert the frame to a format that Tkinter can display
        img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(img)
        imgtk = ImageTk.PhotoImage(image=img)
        frame_label.imgtk = imgtk
        frame_label.configure(image=imgtk)

        if end_attendance_status:
            end_attendance()
        else:
            update_id = frame_label.after(10, update_frame)  # Menyimpan ID pemanggilan update_frame

# Function to start attendance
def start_attendance():
    global start_attendance_time
    start_attendance_time = datetime.now().strftime("%A, %d %B %Y %H:%M:%S")
    global attendance_running
    attendance_running = True
    update_frame()
    start_button.pack_forget()
    end_button.pack(side="top", padx=10, pady=10)

def end_attendance():
    global end_attendance_time
    end_attendance_time = datetime.now().strftime("%A, %d %B %Y %H:%M:%S")
    global attendance_running
    attendance_running = False

    frame_label.after_cancel(update_id)  # Menghentikan pemanggilan update_frame
    start_button.pack_forget()  # Menghilangkan tombol "Start Attendance
    end_button.pack_forget()  # Menghilangkan tombol "End Attendance"
    save_button.pack(side="top", padx=10, pady=5)  # Menampilkan tombol "Save"
    export_button.pack(side="top", padx=10, pady=5)  # Menampilkan tombol "Export"

# Create "Start Attendance" button
start_button = Button(root, text="Start Attendance", command=start_attendance, bg="green", fg="white", width=20)
start_button.pack(side="top", padx=10, pady=10)

# Create "End Attendance" button
end_button = Button(root, text="End Attendance", command=end_attendance, bg="red", fg="white", width=20)

def save_attendance() :
    # Create JSON data for attendance
    attendance_data = {}
    selected_class = class_var.get()
    if selected_class != "Select Class":

        students = class_data[selected_class]["student"]
        student_attendance = []
        for student in students:
            npm = student["npm"]
            name = student["name"]
            status = "Alpha"
            timestamp = "-"
            # Find the latest status and timestamp displayed in student_table
            for row_id in student_table.get_children():
                row_values = student_table.item(row_id)["values"]
                # Convert npm to string for comparison
                npm_str = str(npm)
                if str(row_values[1]) == npm_str:  # Check if the current row corresponds to the student
                    # Always update status and timestamp
                    status = row_values[3]
                    timestamp = row_values[4]
                    break  # Once found, no need to continue looping
            student_attendance.append({"npm": npm_str, "name": name, "status": status, "timestamp": timestamp})

        attendance_data[selected_class] = {
            "Kode_MK": class_data[selected_class]["Kode_MK"],
            "Nama_MK": selected_class,
            "Kelas": class_data[selected_class]["Kelas"],
            "Nama_Kelas": class_data[selected_class]["Kelas"],
            "Kode": class_data[selected_class]["Kode"],
            "Nama_Dosen": class_data[selected_class]["Nama_Dosen"],
            "No_Urut": class_data[selected_class]["No_Urut"],
            "Pertemuan_Ke": class_data[selected_class]["Pertemuan_Ke"],
            "Tanggal": datetime.now().strftime("%d-%m-%Y"),
            "Hari": class_data[selected_class]["Hari"],
            "Waktu": class_data[selected_class]["Waktu"],
            "Ruang": class_data[selected_class]["Ruang"],
            "Dosen_Pengajar": class_data[selected_class]["Kode"],
            "Nama": class_data[selected_class]["Nama_Dosen"],
            "RPS": class_data[selected_class]["RPS"],
            "Alasan": alasan_entry.get(),  # Mengambil nilai dari entry alasan
            "Evaluasi_Pembelajaran": evaluasi_entry.get(),  # Mengambil nilai dari entry evaluasi
            "Mulai_Absensi": start_attendance_time,
            "Akhir_Absensi": end_attendance_time,
            "student": student_attendance
        }

        # Save attendance data to JSON file
        with open("attendance.json", "w") as json_file:
            json.dump(attendance_data, json_file, indent=4)
        
        # root.destroy()  # Menutup jendela setelah data tersimpan
        messagebox.showinfo("Success", "Berhasil Menyimpan Absensi")

def export_attendance():
    selected_class = class_var.get()
    if selected_class != "Select Class":
        # Prepare data for exporting to Excel
        students = class_data[selected_class]["student"]
        attendance_excel_data = []
        for student in students:
            npm = student["npm"]
            name = student["name"]
            status = "Alpha"
            timestamp = "-"
            # Find the latest status and timestamp displayed in student_table
            for row_id in student_table.get_children():
                row_values = student_table.item(row_id)["values"]
                # Convert npm to string for comparison
                npm_str = str(npm)
                if str(row_values[1]) == npm_str:  # Check if the current row corresponds to the student
                    # Always update status and timestamp
                    status = row_values[3]
                    timestamp = row_values[4]
                    break  # Once found, no need to continue looping
            attendance_excel_data.append([npm_str, name, status, timestamp])

        # Get additional data displayed in load_and_display_data
        additional_data = {
            "Kode MK": class_data[selected_class]["Kode_MK"],
            "Nama MK": selected_class,
            "Kelas": class_data[selected_class]["Kelas"],
            "Nama Kelas": class_data[selected_class]["Kelas"],
            "Kode": class_data[selected_class]["Kode"],
            "Nama Dosen": class_data[selected_class]["Nama_Dosen"],
            "No. Urut": class_data[selected_class]["No_Urut"],
            "Pertemuan Ke": class_data[selected_class]["Pertemuan_Ke"],
            "Tanggal": datetime.now().strftime("%d-%m-%Y"),
            "Hari": class_data[selected_class]["Hari"],
            "Waktu": class_data[selected_class]["Waktu"],
            "Ruang": class_data[selected_class]["Ruang"],
            "Dosen Pengajar": class_data[selected_class]["Kode"],
            "Nama": class_data[selected_class]["Nama_Dosen"],
            "RPS": class_data[selected_class]["RPS"],
            "Alasan": alasan_entry.get(),  # Mengambil nilai dari entry alasan
            "Evaluasi Pembelajaran": evaluasi_entry.get(),  # Mengambil nilai dari entry evaluasi
            "Mulai Absensi": start_attendance_time,
            "Akhir Absensi": end_attendance_time
        }

        # Export data to Excel file
        wb = Workbook()
        ws = wb.active

        # Tabel 1: Attribute, Value
        ws.append(["Attribute", "Value"])  # Header untuk data tambahan
        # Set header alignment to center
        for cell in ws["A1:B1"]:
            for c in cell:
                c.alignment = Alignment(horizontal="center")
        for key, value in additional_data.items():
            ws.append([key, value])  # Menambahkan data tambahan ke file Excel

        # Apply border to table 1
        for row in ws.iter_rows(min_row=1, max_row=len(additional_data)+1, min_col=1, max_col=2):
            for cell in row:
                cell.border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

        # Add a blank row between tables
        ws.append([])

        # Tabel 2: NPM, Name, Status, Timestamp
        ws.append(["NPM", "Name", "Status", "Timestamp"])  # Header untuk data mahasiswa
        # Set header alignment to center
        for cell in ws["A" + str(len(additional_data) + 3):"D" + str(len(additional_data) + 3)]:
            for c in cell:
                c.alignment = Alignment(horizontal="center")
        for data_row in attendance_excel_data:
            ws.append(data_row)  # Menambahkan data mahasiswa ke file Excel

        # Apply border to table 2
        for row in ws.iter_rows(min_row=len(additional_data) + 3, min_col=1, max_col=4):
            for cell in row:
                cell.border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20

        # Membuat nama file yang sesuai dengan kelas dan timestamp saat ini
        timestampdate = datetime.now().strftime("%d-%m-%Y")
        file_name = f"attendance_{selected_class}_{timestampdate}.xlsx"
        folder_name = "recap_attendance"
        # Membuat folder "recap_attendance" jika belum ada
        if not os.path.exists(folder_name):
            os.makedirs(folder_name)
        file_path = os.path.join(folder_name, file_name)

        wb.save(file_path)

        # root.destroy()  # Menutup jendela setelah data tersimpan
        messagebox.showinfo("Success", "Berhasil Export Absensi")

# Create "Save Attendance" button
save_button = Button(root, text="Save", command=save_attendance, bg="blue", fg="white", width=20)

# Create "Export Attendance" button
export_button = Button(root, text="Export", command=export_attendance, bg="green", fg="white", width=20)

# Run the Tkinter event loop
root.mainloop()